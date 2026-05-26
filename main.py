
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from pathlib import Path
from datetime import datetime
import os
import re
import sqlite3
import hashlib
import shutil

try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

APP_VERSION = "V1.0.0.9_BASE_FIND_FLEX"
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)

SQLITE_DB = BASE_DIR / "dsystem_ar_api.db"
BASE_XLSX_CACHE = BASE_DIR / "base_xlsx_cache.xlsx"
HEADER_ROW = 5
START_COL = 2

app = FastAPI(
    title="DSYSTEM AR API",
    version=APP_VERSION,
    description="API central do DSYSTEM AR Scanner Mobile e AR Painel com suporte PostgreSQL."
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class LoginRequest(BaseModel):
    usuario: str
    senha: str

class CreateUserRequest(BaseModel):
    usuario: str
    nome: str
    senha: str
    perfil: str = "operador"

class ChangePasswordRequest(BaseModel):
    usuario: str
    senha_atual: str
    nova_senha: str

class AdminChangePasswordRequest(BaseModel):
    usuario: str
    nova_senha: str


def normalize_database_url(url: str) -> str:
    if url.startswith("postgres://"):
        return "postgresql://" + url[len("postgres://"):]
    return url


def get_conn():
    if USE_POSTGRES:
        if psycopg2 is None:
            raise RuntimeError("psycopg2-binary não está instalado.")
        return psycopg2.connect(
            normalize_database_url(DATABASE_URL),
            cursor_factory=psycopg2.extras.RealDictCursor
        )

    conn = sqlite3.connect(SQLITE_DB)
    conn.row_factory = sqlite3.Row
    return conn


def q(sql: str) -> str:
    """Converte placeholders ? para %s quando estiver usando PostgreSQL."""
    return sql.replace("?", "%s") if USE_POSTGRES else sql


def rows_to_dicts(rows):
    return [dict(r) for r in rows]


def row_to_dict(row):
    return dict(row) if row else None


def fetchone(sql, params=()):
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(q(sql), params)
        row = cur.fetchone()
        return row_to_dict(row)
    finally:
        conn.close()


def fetchall(sql, params=()):
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(q(sql), params)
        rows = cur.fetchall()
        return rows_to_dicts(rows)
    finally:
        conn.close()


def execute(sql, params=()):
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(q(sql), params)
        conn.commit()
        return cur
    finally:
        conn.close()


def init_db():
    conn = get_conn()
    try:
        cur = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    usuario TEXT UNIQUE NOT NULL,
                    nome TEXT NOT NULL,
                    senha TEXT NOT NULL,
                    perfil TEXT NOT NULL DEFAULT 'operador',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS ars (
                    id SERIAL PRIMARY KEY,
                    created_at TEXT NOT NULL,
                    original_name TEXT,
                    file_name TEXT NOT NULL,
                    file_path TEXT,
                    file_data BYTEA,
                    mime_type TEXT DEFAULT 'application/pdf',
                    instalacao TEXT,
                    medidor TEXT,
                    nome_cliente TEXT,
                    status TEXT DEFAULT 'Pendente',
                    origem TEXT DEFAULT 'api',
                    observacao TEXT,
                    operador_usuario TEXT,
                    operador_nome TEXT,
                    operador_perfil TEXT
                )
            """)


            cur.execute("""
                CREATE TABLE IF NOT EXISTS base_xlsx (
                    id SERIAL PRIMARY KEY,
                    base_key TEXT UNIQUE NOT NULL,
                    aba TEXT,
                    instalacao TEXT,
                    instalacao_norm TEXT,
                    medidor TEXT,
                    medidor_norm TEXT,
                    nome_cliente TEXT,
                    nome_norm TEXT,
                    instalacao_compact TEXT,
                    medidor_compact TEXT,
                    nome_compact TEXT,
                    import_batch TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)


            for col, typ in [
                ("instalacao_compact", "TEXT"),
                ("medidor_compact", "TEXT"),
                ("nome_compact", "TEXT"),
            ]:
                cur.execute("""
                    SELECT column_name FROM information_schema.columns
                    WHERE table_name='base_xlsx' AND column_name=%s
                """, (col,))
                if not cur.fetchone():
                    cur.execute(f"ALTER TABLE base_xlsx ADD COLUMN {col} {typ}")

            # Migrações defensivas para bases antigas
            for col, typ in [
                ("file_data", "BYTEA"),
                ("mime_type", "TEXT DEFAULT 'application/pdf'"),
                ("operador_usuario", "TEXT"),
                ("operador_nome", "TEXT"),
                ("operador_perfil", "TEXT"),
                ("instalacao_compact", "TEXT"),
                ("medidor_compact", "TEXT"),
                ("nome_compact", "TEXT"),
            ]:
                cur.execute("""
                    SELECT column_name FROM information_schema.columns
                    WHERE table_name='ars' AND column_name=%s
                """, (col,))
                if not cur.fetchone():
                    cur.execute(f"ALTER TABLE ars ADD COLUMN {col} {typ}")

        else:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    usuario TEXT UNIQUE NOT NULL,
                    nome TEXT NOT NULL,
                    senha TEXT NOT NULL,
                    perfil TEXT NOT NULL DEFAULT 'operador',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS ars (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    original_name TEXT,
                    file_name TEXT NOT NULL,
                    file_path TEXT,
                    file_data BLOB,
                    mime_type TEXT DEFAULT 'application/pdf',
                    instalacao TEXT,
                    medidor TEXT,
                    nome_cliente TEXT,
                    status TEXT DEFAULT 'Pendente',
                    origem TEXT DEFAULT 'api',
                    observacao TEXT,
                    operador_usuario TEXT,
                    operador_nome TEXT,
                    operador_perfil TEXT
                )
            """)


            cur.execute("""
                CREATE TABLE IF NOT EXISTS base_xlsx (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    base_key TEXT UNIQUE NOT NULL,
                    aba TEXT,
                    instalacao TEXT,
                    instalacao_norm TEXT,
                    medidor TEXT,
                    medidor_norm TEXT,
                    nome_cliente TEXT,
                    nome_norm TEXT,
                    instalacao_compact TEXT,
                    medidor_compact TEXT,
                    nome_compact TEXT,
                    import_batch TEXT,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cols = [r["name"] for r in cur.execute("PRAGMA table_info(ars)").fetchall()]
            migrations = {
                "file_data": "ALTER TABLE ars ADD COLUMN file_data BLOB",
                "mime_type": "ALTER TABLE ars ADD COLUMN mime_type TEXT DEFAULT 'application/pdf'",
                "operador_usuario": "ALTER TABLE ars ADD COLUMN operador_usuario TEXT",
                "operador_nome": "ALTER TABLE ars ADD COLUMN operador_nome TEXT",
                "operador_perfil": "ALTER TABLE ars ADD COLUMN operador_perfil TEXT",
            }
            for col, sql in migrations.items():
                if col not in cols:
                    cur.execute(sql)

            cols_base_xlsx = [r["name"] for r in cur.execute("PRAGMA table_info(base_xlsx)").fetchall()]
            base_migrations = {
                "instalacao_compact": "ALTER TABLE base_xlsx ADD COLUMN instalacao_compact TEXT",
                "medidor_compact": "ALTER TABLE base_xlsx ADD COLUMN medidor_compact TEXT",
                "nome_compact": "ALTER TABLE base_xlsx ADD COLUMN nome_compact TEXT",
            }
            for col, sql in base_migrations.items():
                if col not in cols_base_xlsx:
                    cur.execute(sql)

        cur.execute(q("SELECT id FROM users WHERE usuario=?"), ("admin",))
        if not cur.fetchone():
            cur.execute(
                q("INSERT INTO users(usuario,nome,senha,perfil) VALUES(?,?,?,?)"),
                ("admin", "Administrador", "admin123", "admin")
            )

        conn.commit()
    finally:
        conn.close()


def sanitize_filename(name: str) -> str:
    name = (name or "").strip().replace(" ", "_")
    name = re.sub(r"[^A-Za-z0-9_.\-]", "", name)
    return name or "arquivo.pdf"


def media_type_from_ext(ext: str) -> str:
    ext = (ext or "").lower()
    if ext == ".pdf":
        return "application/pdf"
    if ext in [".jpg", ".jpeg"]:
        return "image/jpeg"
    if ext == ".png":
        return "image/png"
    if ext == ".webp":
        return "image/webp"
    return "application/octet-stream"



def norm_header(v):
    s = str(v or "").strip().upper()
    mapa = {"Á":"A","À":"A","Â":"A","Ã":"A","É":"E","Ê":"E","Í":"I","Ó":"O","Õ":"O","Ô":"O","Ú":"U","Ç":"C"}
    for a,b in mapa.items():
        s = s.replace(a,b)
    return re.sub(r"\s+", " ", s)

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def norm_key(v):
    return re.sub(r"\s+", " ", str(v or "").strip().upper())

def norm_lookup(v):
    # Busca flexível: remove espaços, pontuação, barras, hífens e zeros/formatos estranhos.
    return re.sub(r"[^A-Z0-9]", "", norm_key(v))

def only_digits(v):
    return re.sub(r"\D+", "", str(v or ""))

def base_unique_key(aba, instalacao, medidor, nome_cliente):
    raw = "|".join([norm_key(aba), norm_key(instalacao), norm_key(medidor), norm_key(nome_cliente)])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()

def auto_col(headers_norm, names):
    for target in names:
        t = norm_header(target)
        for i,h in enumerate(headers_norm):
            if h == t or t in h:
                return i + 1
    return None

def load_xlsx_book_readonly():
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(500, detail="openpyxl não instalado")
    if not BASE_XLSX_CACHE.exists():
        raise HTTPException(404, detail="Nenhuma base XLSX importada ainda")
    return load_workbook(BASE_XLSX_CACHE, data_only=True, read_only=True)


def build_name(instalacao: str = "", medidor: str = "", ext: str = ".pdf") -> str:
    instalacao = (instalacao or "").strip()
    medidor = (medidor or "").strip()

    if instalacao or medidor:
        return f"AR_CARTACONVITE_INST_{instalacao}_MD_{medidor}{ext}"

    row = fetchone("SELECT COUNT(*) AS total FROM ars WHERE file_name LIKE ?", ("AR_CARTACONVITE_PAG.%",))
    total = (row or {}).get("total", 0) or 0
    return f"AR_CARTACONVITE_PAG.{int(total) + 1:02d}{ext}"


init_db()


@app.get("/")
def root():
    return {
        "app": "DSYSTEM AR API",
        "version": APP_VERSION,
        "status": "online",
        "database": "postgresql" if USE_POSTGRES else "sqlite",
        "storage": "database_file_data",
        "endpoints": [
            "/api/status",
            "/api/login",
            "/api/users",
            "/api/ars",
            "/api/upload",
            "/api/ars/{id}/download",
            "/api/ars/{id}/view",
            "/api/upload-base-cache",
            "/api/import-base",
            "/api/base/find"
        ]
    }


@app.get("/api/status")
def status():
    row = fetchone("SELECT COUNT(*) AS total FROM ars")
    users = fetchone("SELECT COUNT(*) AS total FROM users")
    return {
        "status": "online",
        "version": APP_VERSION,
        "database": "postgresql" if USE_POSTGRES else "sqlite",
        "total_ars": int((row or {}).get("total", 0) or 0),
        "total_users": int((users or {}).get("total", 0) or 0)
    }


@app.post("/api/login")
def login(data: LoginRequest):
    user = fetchone(
        "SELECT id, usuario, nome, perfil FROM users WHERE usuario=? AND senha=?",
        (data.usuario, data.senha)
    )
    if not user:
        raise HTTPException(status_code=401, detail="Usuário ou senha inválidos")
    return user


@app.get("/api/users")
def list_users():
    return fetchall("SELECT id, usuario, nome, perfil FROM users ORDER BY id DESC")


@app.post("/api/users")
def create_user(data: CreateUserRequest):
    usuario = (data.usuario or "").strip()
    nome = (data.nome or "").strip()
    senha = data.senha or ""
    perfil = (data.perfil or "operador").strip().lower()

    if perfil not in ["admin", "operador"]:
        perfil = "operador"

    if not usuario or not nome or not senha:
        raise HTTPException(status_code=400, detail="Preencha usuário, nome e senha")

    exists = fetchone("SELECT id FROM users WHERE usuario=?", (usuario,))
    if exists:
        raise HTTPException(status_code=400, detail="Usuário já existe")

    execute(
        "INSERT INTO users(usuario,nome,senha,perfil) VALUES(?,?,?,?)",
        (usuario, nome, senha, perfil)
    )
    return {"success": True}


@app.put("/api/change-password")
def change_password(data: ChangePasswordRequest):
    user = fetchone(
        "SELECT id FROM users WHERE usuario=? AND senha=?",
        (data.usuario, data.senha_atual)
    )
    if not user:
        raise HTTPException(status_code=401, detail="Senha atual inválida")

    execute("UPDATE users SET senha=? WHERE usuario=?", (data.nova_senha, data.usuario))
    return {"success": True}


@app.put("/api/admin/change-password")
def admin_change_password(data: AdminChangePasswordRequest):
    if not data.usuario or not data.nova_senha:
        raise HTTPException(status_code=400, detail="Informe usuário e nova senha")

    user = fetchone("SELECT id FROM users WHERE usuario=?", (data.usuario,))
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    execute("UPDATE users SET senha=? WHERE usuario=?", (data.nova_senha, data.usuario))
    return {"success": True}


@app.delete("/api/users/{usuario}")
def delete_user(usuario: str):
    if usuario == "admin":
        raise HTTPException(status_code=400, detail="Admin padrão não pode ser removido")

    execute("DELETE FROM users WHERE usuario=?", (usuario,))
    return {"success": True}


@app.post("/api/upload")
async def upload_ar(
    file: UploadFile = File(...),
    instalacao: str = Form(""),
    medidor: str = Form(""),
    nome_cliente: str = Form(""),
    origem: str = Form("mobile"),
    operador_usuario: str = Form(""),
    operador_nome: str = Form(""),
    operador_perfil: str = Form("")
):
    original = file.filename or "arquivo.pdf"
    ext = Path(original).suffix.lower() or ".pdf"

    if ext not in [".pdf", ".jpg", ".jpeg", ".png", ".webp"]:
        raise HTTPException(status_code=400, detail="Formato não permitido.")

    final_name = sanitize_filename(build_name(instalacao, medidor, ext))
    mime_type = media_type_from_ext(ext)
    file_bytes = await file.read()

    # Mantém cópia local quando possível, mas a fonte confiável é file_data no banco.
    target = UPLOAD_DIR / final_name
    try:
        if target.exists():
            stem = target.stem
            suffix = target.suffix
            n = 2
            while (UPLOAD_DIR / f"{stem}_{n}{suffix}").exists():
                n += 1
            target = UPLOAD_DIR / f"{stem}_{n}{suffix}"
            final_name = target.name
        target.write_bytes(file_bytes)
        file_path = str(target)
    except Exception:
        file_path = ""

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    try:
        cur = conn.cursor()

        if USE_POSTGRES:
            cur.execute("""
                INSERT INTO ars (
                    created_at, original_name, file_name, file_path, file_data, mime_type,
                    instalacao, medidor, nome_cliente, status, origem,
                    operador_usuario, operador_nome, operador_perfil
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                now, original, final_name, file_path, psycopg2.Binary(file_bytes), mime_type,
                instalacao.strip() or None,
                medidor.strip() or None,
                nome_cliente.strip() or None,
                "Pendente",
                origem,
                operador_usuario.strip() or None,
                operador_nome.strip() or None,
                operador_perfil.strip() or None
            ))
            new_id = cur.fetchone()["id"]
        else:
            cur.execute("""
                INSERT INTO ars (
                    created_at, original_name, file_name, file_path, file_data, mime_type,
                    instalacao, medidor, nome_cliente, status, origem,
                    operador_usuario, operador_nome, operador_perfil
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                now, original, final_name, file_path, file_bytes, mime_type,
                instalacao.strip() or None,
                medidor.strip() or None,
                nome_cliente.strip() or None,
                "Pendente",
                origem,
                operador_usuario.strip() or None,
                operador_nome.strip() or None,
                operador_perfil.strip() or None
            ))
            new_id = cur.lastrowid

        conn.commit()
    finally:
        conn.close()

    return {
        "ok": True,
        "id": new_id,
        "file_name": final_name,
        "instalacao": instalacao,
        "medidor": medidor,
        "nome_cliente": nome_cliente,
        "operador_usuario": operador_usuario,
        "operador_nome": operador_nome,
        "operador_perfil": operador_perfil,
        "status": "Pendente"
    }


@app.get("/api/ars")
def list_ars(
    instalacao: str = "",
    medidor: str = "",
    nome_cliente: str = "",
    status: str = ""
):
    sql = """
        SELECT id, created_at, original_name, file_name, file_path, mime_type,
               instalacao, medidor, nome_cliente, status, origem, observacao,
               operador_usuario, operador_nome, operador_perfil
        FROM ars
        WHERE 1=1
    """
    params = []

    like_op = "ILIKE" if USE_POSTGRES else "LIKE"

    if instalacao:
        sql += f" AND instalacao {like_op} ?"
        params.append(f"%{instalacao}%")
    if medidor:
        sql += f" AND medidor {like_op} ?"
        params.append(f"%{medidor}%")
    if nome_cliente:
        sql += f" AND nome_cliente {like_op} ?"
        params.append(f"%{nome_cliente}%")
    if status:
        sql += " AND status = ?"
        params.append(status)

    sql += " ORDER BY id DESC"
    return fetchall(sql, tuple(params))


@app.get("/api/ars/{ar_id}")
def get_ar(ar_id: int):
    row = fetchone("""
        SELECT id, created_at, original_name, file_name, file_path, mime_type,
               instalacao, medidor, nome_cliente, status, origem, observacao,
               operador_usuario, operador_nome, operador_perfil
        FROM ars
        WHERE id=?
    """, (ar_id,))
    if not row:
        raise HTTPException(status_code=404, detail="AR não encontrado.")
    return row


@app.put("/api/ars/{ar_id}")
def update_ar(
    ar_id: int,
    instalacao: str = Form(""),
    medidor: str = Form(""),
    nome_cliente: str = Form(""),
    status: str = Form("Pendente"),
    observacao: str = Form(""),
    operador_usuario: str = Form(""),
    operador_nome: str = Form(""),
    operador_perfil: str = Form("")
):
    row = fetchone("SELECT * FROM ars WHERE id=?", (ar_id,))
    if not row:
        raise HTTPException(status_code=404, detail="AR não encontrado.")

    old_name = row.get("file_name") or "arquivo.pdf"
    ext = Path(old_name).suffix or ".pdf"
    new_name = sanitize_filename(build_name(instalacao, medidor, ext))

    execute("""
        UPDATE ars
        SET instalacao=?, medidor=?, nome_cliente=?, status=?, observacao=?,
            file_name=?,
            operador_usuario=COALESCE(NULLIF(?, ''), operador_usuario),
            operador_nome=COALESCE(NULLIF(?, ''), operador_nome),
            operador_perfil=COALESCE(NULLIF(?, ''), operador_perfil)
        WHERE id=?
    """, (
        instalacao.strip() or None,
        medidor.strip() or None,
        nome_cliente.strip() or None,
        status,
        observacao.strip() or None,
        new_name,
        operador_usuario.strip(),
        operador_nome.strip(),
        operador_perfil.strip(),
        ar_id
    ))

    return {"ok": True, "id": ar_id, "file_name": new_name, "status": status}


@app.delete("/api/ars/{ar_id}")
def delete_ar(ar_id: int):
    row = fetchone("SELECT * FROM ars WHERE id=?", (ar_id,))
    if not row:
        raise HTTPException(status_code=404, detail="AR não encontrado.")

    path = Path(row.get("file_path") or "")
    try:
        if path.exists():
            path.unlink()
    except Exception:
        pass

    execute("DELETE FROM ars WHERE id=?", (ar_id,))
    return {"ok": True, "deleted": ar_id}


def file_response_from_row(row):
    if not row:
        raise HTTPException(status_code=404, detail="AR não encontrado.")

    filename = row.get("file_name") or "arquivo.pdf"
    mime_type = row.get("mime_type") or media_type_from_ext(Path(filename).suffix)

    file_data = row.get("file_data")
    if file_data is not None:
        if isinstance(file_data, memoryview):
            file_data = file_data.tobytes()
        return Response(
            content=bytes(file_data),
            media_type=mime_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    path = Path(row.get("file_path") or "")
    if path.exists():
        return FileResponse(path, filename=filename, media_type=mime_type)

    raise HTTPException(status_code=404, detail="Arquivo não encontrado.")


@app.get("/api/ars/{ar_id}/download")
def download_ar(ar_id: int):
    row = fetchone("SELECT * FROM ars WHERE id=?", (ar_id,))
    return file_response_from_row(row)


@app.get("/api/download/{ar_id}")
def download_ar_alias(ar_id: int):
    row = fetchone("SELECT * FROM ars WHERE id=?", (ar_id,))
    return file_response_from_row(row)


@app.get("/api/ars/{ar_id}/view")
def view_ar(ar_id: int):
    row = fetchone("SELECT * FROM ars WHERE id=?", (ar_id,))
    if not row:
        raise HTTPException(status_code=404, detail="AR não encontrado.")

    filename = row.get("file_name") or "arquivo.pdf"
    mime_type = row.get("mime_type") or media_type_from_ext(Path(filename).suffix)
    file_data = row.get("file_data")

    if file_data is not None:
        if isinstance(file_data, memoryview):
            file_data = file_data.tobytes()
        return Response(content=bytes(file_data), media_type=mime_type)

    path = Path(row.get("file_path") or "")
    if path.exists():
        return FileResponse(path, media_type=mime_type)

    raise HTTPException(status_code=404, detail="Arquivo não encontrado.")



@app.post("/api/upload-base-cache")
async def upload_base_cache(file: UploadFile = File(...)):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo XLSX")

    with BASE_XLSX_CACHE.open("wb") as out:
        shutil.copyfileobj(file.file, out)

    wb = load_xlsx_book_readonly()
    try:
        sheets = list(wb.sheetnames)
    finally:
        wb.close()

    return {"ok": True, "sheets": sheets}


@app.get("/api/base/sheets")
def base_sheets():
    wb = load_xlsx_book_readonly()
    try:
        return {"sheets": list(wb.sheetnames)}
    finally:
        wb.close()


@app.get("/api/base/columns")
def base_columns(sheet: str):
    wb = load_xlsx_book_readonly()
    try:
        if sheet not in wb.sheetnames:
            raise HTTPException(404, detail="Aba não encontrada")
        ws = wb[sheet]
        headers = [cell_str(ws.cell(HEADER_ROW, col).value) for col in range(START_COL, ws.max_column + 1)]
        return {"columns": headers}
    finally:
        wb.close()


@app.post("/api/import-base")
def import_base(
    sheet: str = Form("TODAS"),
    col_instalacao: str = Form(""),
    col_medidor: str = Form(""),
    col_nome_cliente: str = Form("")
):
    wb = load_xlsx_book_readonly()
    batch = datetime.now().strftime("%Y%m%d%H%M%S")
    count_insert = 0
    count_update = 0
    count_seen = 0

    try:
        sheets = list(wb.sheetnames) if sheet == "TODAS" else [sheet]
        conn = get_conn()
        try:
            cur = conn.cursor()
            for sh in sheets:
                if sh not in wb.sheetnames:
                    continue

                ws = wb[sh]
                headers_raw = [cell_str(ws.cell(HEADER_ROW, col).value) for col in range(START_COL, ws.max_column + 1)]
                headers_norm = [norm_header(h) for h in headers_raw]

                def col_by_selected(sel, auto_names):
                    if sel:
                        try:
                            return headers_raw.index(sel) + START_COL
                        except ValueError:
                            pass
                    idx = auto_col(headers_norm, auto_names)
                    return (idx + START_COL - 1) if idx else None

                col_i = col_by_selected(col_instalacao, ["INSTALACAO","INST","UC","UNIDADE CONSUMIDORA","INSTALAÇÃO"])
                col_m = col_by_selected(col_medidor, ["MEDIDOR","MD","N MEDIDOR","NUMERO MEDIDOR","SERIAL","N DE SERIE","Nº MEDIDOR"])
                col_n = col_by_selected(col_nome_cliente, ["NOME_CLIENTE","NOME CLIENTE","CLIENTE","NOME","NOME DO CLIENTE"])

                for row in range(HEADER_ROW + 1, ws.max_row + 1):
                    inst = cell_str(ws.cell(row, col_i).value) if col_i else ""
                    md = cell_str(ws.cell(row, col_m).value) if col_m else ""
                    nome = cell_str(ws.cell(row, col_n).value) if col_n else ""

                    if not (inst or md or nome):
                        continue

                    bkey = base_unique_key(sh, inst, md, nome)
                    inst_n = norm_key(inst)
                    md_n = norm_key(md)
                    nome_n = norm_key(nome)
                    inst_c = norm_lookup(inst)
                    md_c = norm_lookup(md)
                    nome_c = norm_lookup(nome)
                    count_seen += 1

                    exists_sql = "SELECT id FROM base_xlsx WHERE base_key=?"
                    cur.execute(q(exists_sql), (bkey,))
                    exists = cur.fetchone()

                    if exists:
                        cur.execute(q("""
                            UPDATE base_xlsx
                            SET aba=?, instalacao=?, instalacao_norm=?, medidor=?, medidor_norm=?,
                                nome_cliente=?, nome_norm=?, instalacao_compact=?, medidor_compact=?, nome_compact=?,
                                import_batch=?, updated_at=CURRENT_TIMESTAMP
                            WHERE base_key=?
                        """), (sh, inst, inst_n, md, md_n, nome, nome_n, inst_c, md_c, nome_c, batch, bkey))
                        count_update += 1
                    else:
                        cur.execute(q("""
                            INSERT INTO base_xlsx(
                                base_key, aba, instalacao, instalacao_norm,
                                medidor, medidor_norm, nome_cliente, nome_norm,
                                instalacao_compact, medidor_compact, nome_compact, import_batch
                            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                        """), (bkey, sh, inst, inst_n, md, md_n, nome, nome_n, inst_c, md_c, nome_c, batch))
                        count_insert += 1

            conn.commit()
        finally:
            conn.close()

    finally:
        wb.close()

    return {
        "ok": True,
        "count": count_seen,
        "inserted": count_insert,
        "updated": count_update,
        "sheets": sheets
    }


@app.get("/api/base/find")
def find_base(instalacao: str = "", medidor: str = "", nome_cliente: str = ""):
    inst = norm_key(instalacao)
    md = norm_key(medidor)
    nome = norm_key(nome_cliente)

    inst_c = norm_lookup(instalacao)
    md_c = norm_lookup(medidor)
    nome_c = norm_lookup(nome_cliente)

    row = None

    # 1) Melhor caso: instalação + medidor, flexível.
    if inst_c and md_c:
        row = fetchone("""
            SELECT * FROM base_xlsx
            WHERE instalacao_compact=? AND medidor_compact=?
            ORDER BY updated_at DESC, id DESC LIMIT 1
        """, (inst_c, md_c))

    # 2) Medidor exato/compacto geralmente é a chave mais confiável.
    if not row and md_c:
        row = fetchone("""
            SELECT * FROM base_xlsx
            WHERE medidor_compact=?
            ORDER BY updated_at DESC, id DESC LIMIT 1
        """, (md_c,))

    # 3) Instalação exata/compacta.
    if not row and inst_c:
        row = fetchone("""
            SELECT * FROM base_xlsx
            WHERE instalacao_compact=?
            ORDER BY updated_at DESC, id DESC LIMIT 1
        """, (inst_c,))

    # 4) Fallback parcial com LIKE/ILIKE para casos com zeros, hífen ou sufixos.
    like_op = "ILIKE" if USE_POSTGRES else "LIKE"

    if not row and md_c and len(md_c) >= 4:
        rows = fetchall(f"""
            SELECT * FROM base_xlsx
            WHERE medidor_compact {like_op} ?
            ORDER BY updated_at DESC, id DESC LIMIT 1
        """, (f"%{md_c}%",))
        row = rows[0] if rows else None

    if not row and inst_c and len(inst_c) >= 4:
        rows = fetchall(f"""
            SELECT * FROM base_xlsx
            WHERE instalacao_compact {like_op} ?
            ORDER BY updated_at DESC, id DESC LIMIT 1
        """, (f"%{inst_c}%",))
        row = rows[0] if rows else None

    if not row and nome_c and len(nome_c) >= 3:
        rows = fetchall(f"""
            SELECT * FROM base_xlsx
            WHERE nome_compact {like_op} ?
            ORDER BY updated_at DESC, id DESC LIMIT 1
        """, (f"%{nome_c}%",))
        row = rows[0] if rows else None

    return row or {}



@app.post("/api/base/reindex")
def base_reindex():
    rows = fetchall("SELECT id, instalacao, medidor, nome_cliente FROM base_xlsx")
    conn = get_conn()
    try:
        cur = conn.cursor()
        for r in rows:
            cur.execute(q("""
                UPDATE base_xlsx
                SET instalacao_compact=?, medidor_compact=?, nome_compact=?
                WHERE id=?
            """), (
                norm_lookup(r.get("instalacao")),
                norm_lookup(r.get("medidor")),
                norm_lookup(r.get("nome_cliente")),
                r.get("id")
            ))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "updated": len(rows)}


@app.get("/api/base/stats")
def base_stats():
    row = fetchone("SELECT COUNT(*) AS total FROM base_xlsx")
    return {"ok": True, "total": int((row or {}).get("total", 0) or 0)}


@app.get("/TESTE_UPLOAD.html")
def teste_upload_html():
    teste = BASE_DIR / "TESTE_UPLOAD.html"
    if not teste.exists():
        raise HTTPException(status_code=404, detail="TESTE_UPLOAD.html não encontrado.")
    return FileResponse(teste, media_type="text/html")
